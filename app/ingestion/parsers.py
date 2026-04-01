"""Multi-format document parser.

Handles PDF, DOCX, HTML, CSV, TXT, MD, XLSX, XLS formats.
If extracted text is <50 characters, flags needs_ocr=True.
"""
import os
import tempfile
import logging
from pathlib import Path
from typing import Dict, Any

logger = logging.getLogger(__name__)

# Lazy-load heavy dependencies
_pdf_loader = None
_docx_loader = None
_bs4_loader = None
_csv_loader = None
_text_loader = None


def _get_pypdf_loader():
    global _pdf_loader
    if _pdf_loader is None:
        from langchain_community.document_loaders import PyPDFLoader
        _pdf_loader = PyPDFLoader
    return _pdf_loader


def _get_docx_loader():
    global _docx_loader
    if _docx_loader is None:
        from langchain_community.document_loaders import Docx2txtLoader
        _docx_loader = Docx2txtLoader
    return _docx_loader


def _get_bs4_loader():
    global _bs4_loader
    if _bs4_loader is None:
        from langchain_community.document_loaders import BS4HTMLLoader
        _bs4_loader = BS4HTMLLoader
    return _bs4_loader


def _get_csv_loader():
    global _csv_loader
    if _csv_loader is None:
        from langchain_community.document_loaders import CSVLoader
        _csv_loader = CSVLoader
    return _csv_loader


def _get_text_loader():
    global _text_loader
    if _text_loader is None:
        from langchain_community.document_loaders import TextLoader
        _text_loader = TextLoader
    return _text_loader


def parse_document(file_path: str) -> Dict[str, Any]:
    """
    Parse a document and extract text content.

    Args:
        file_path: Absolute path to the file.

    Returns:
        {
            "text": str,          # extracted text
            "pages": int,         # number of pages (or 1 for non-paginated)
            "metadata": dict,     # source, filename, extension, etc.
            "needs_ocr": bool,    # True if text is empty/very short
        }
    """
    path = Path(file_path)
    suffix = path.suffix.lower()
    filename = path.name

    metadata = {
        "source": str(file_path),
        "filename": filename,
        "extension": suffix,
    }

    text = ""
    pages = 1

    try:
        if suffix == ".pdf":
            text, pages = _parse_pdf(file_path)
        elif suffix == ".docx":
            text, pages = _parse_docx(file_path)
        elif suffix == ".html":
            text, pages = _parse_html(file_path)
        elif suffix == ".csv":
            text, pages = _parse_csv(file_path)
        elif suffix in (".txt", ".md"):
            text, pages = _parse_text(file_path)
        elif suffix in (".xlsx", ".xls"):
            text, pages = _parse_excel(file_path)
        else:
            logger.warning(f"Unsupported file type: {suffix}")
            return {
                "text": "",
                "pages": 0,
                "metadata": metadata,
                "needs_ocr": False,
            }
    except Exception as e:
        logger.exception(f"Error parsing {file_path}: {e}")
        return {
            "text": "",
            "pages": 0,
            "metadata": metadata,
            "needs_ocr": False,
        }

    # OCR fallback flag — short/no text
    needs_ocr = len(text.strip()) < 50

    return {
        "text": text,
        "pages": pages,
        "metadata": metadata,
        "needs_ocr": needs_ocr,
    }


# ── Per-format parsers ────────────────────────────────────────────────────────


def _parse_pdf(file_path: str) -> tuple[str, int]:
    """Extract text from PDF using PyPDF. Returns (text, page_count)."""
    loader_cls = _get_pypdf_loader()
    loader = loader_cls(file_path)
    docs = loader.load()
    pages = len(docs)
    texts = [doc.page_content for doc in docs]
    return "\n\n".join(texts), pages


def _parse_docx(file_path: str) -> tuple[str, int]:
    """Extract text from DOCX using Docx2txt."""
    loader_cls = _get_docx_loader()
    loader = loader_cls(file_path)
    docs = loader.load()
    # Docx2txt returns a single document
    text = docs[0].page_content if docs else ""
    return text, 1


def _parse_html(file_path: str) -> tuple[str, int]:
    """Extract text from HTML using BeautifulSoup via langchain loader."""
    loader_cls = _get_bs4_loader()
    loader = loader_cls(file_path)
    docs = loader.load()
    texts = [doc.page_content for doc in docs]
    return "\n\n".join(texts), len(docs)


def _parse_csv(file_path: str) -> tuple[str, int]:
    """Extract text from CSV using langchain CSVLoader."""
    loader_cls = _get_csv_loader()
    loader = loader_cls(file_path, encoding="utf-8")
    docs = loader.load()
    texts = [doc.page_content for doc in docs]
    return "\n\n".join(texts), 1


def _parse_text(file_path: str) -> tuple[str, int]:
    """Load plain text or markdown."""
    loader_cls = _get_text_loader()
    loader = loader_cls(file_path, encoding="utf-8")
    docs = loader.load()
    texts = [doc.page_content for doc in docs]
    return "\n\n".join(texts), 1


def _parse_excel(file_path: str) -> tuple[str, int]:
    """Extract text from XLSX/XLS using openpyxl or pandas."""
    suffix = Path(file_path).suffix.lower()

    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheets = wb.sheetnames
        lines = []
        for sheet_name in sheets:
            ws = wb[sheet_name]
            lines.append(f"## Sheet: {sheet_name}")
            for row in ws.iter_rows(values_only=True):
                # Filter out None values
                row_vals = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in row_vals):
                    lines.append("\t".join(row_vals))
        return "\n".join(lines), len(sheets)
    except ImportError:
        logger.warning("openpyxl not available, falling back to pandas for Excel")
        import pandas as pd

        if suffix == ".xlsx":
            dfs = pd.read_excel(file_path, sheet_name=None, engine="openpyxl")
        else:
            dfs = pd.read_excel(file_path, sheet_name=None)
        lines = []
        for sheet_name, df in dfs.items():
            lines.append(f"## Sheet: {sheet_name}")
            lines.append(df.to_string(index=False))
        return "\n".join(lines), len(dfs)
